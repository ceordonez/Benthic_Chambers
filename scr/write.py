#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import pdb
import sys

import pandas as pd
import numpy as np

from openpyxl import load_workbook
from datetime import datetime

def append_df_to_excel(filename, df, bc_name, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            old_res = pd.read_excel(filename, index_col=[0])
            index = old_res.index.values
            if bc_name in index:
                startrow = np.where(index == bc_name)[-1][0] + 1
            else:
                startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    # save the workbook
    writer.save()


def write_excel(path, results, lake, core, meta, exp, new=False):

    names = ['Benthic_Chambers', lake]
    filename = '_'.join(names)
    path_out = os.path.join(path, lake, 'Results', 'Benthic_Chamber')
    if not os.path.exists(path_out):
        os.makedirs(path_out)

    if exp !='':
        resfile = os.path.join(path_out, filename + '_' + exp + '.xlsx')
    else:
        resfile = os.path.join(path_out, filename + '.xlsx')

    if not os.path.isfile(resfile) or new:
        writer = pd.ExcelWriter(resfile, engine='xlsxwriter')
        workbook = writer.book
        for var in results:
            units = {'BC_name':'', 'Start': '', 'End': '',
                     'Fs_lin':'(mmol/m2/d)', 'Fs_dPdt': '(mmol/m2/d)',
                     'Fs_MB': '(mmol/m2/d)', 'k': '(m/d)', 'Cw0': '(umol/l)',
                     'Cwf': '(umol/l)', 'Cwf_dPdt': '(umol/l)', 'Cwf_lin': '(umol/l)',
                     'P': '(hPa)', 'Temp': '(oC)', 'hw': '(cm)', 'ha':'(cm)',
                     'Px_start': '(Pa)', 'Px_end': '(Pa)'}
            worksheet = writer.book.add_worksheet(var[:3])
            bold = workbook.add_format({'bold': True, 'align': 'center'})
            i = 0
            for key in units:
                worksheet.write(0, i, key, bold)
                worksheet.write(1, i, units[key], bold)
                i+=1
            # format excel
            formath = workbook.add_format({'num_format':'0.00'})
            dformat = workbook.add_format({'num_format':'dd/mm/YYYY hh:mm'})
            formath.set_align('center')
            formath.set_align('vcenter')
            dformat.set_align('center')
            dformat.set_align('vcenter')
            worksheet.set_column('A:A', 10, formath)
            worksheet.set_column('B:C', 20, dformat)
            worksheet.set_column('D:F', 13, formath)
            worksheet.set_column('G:R', 10, formath)
        writer.save()
        for var in results:
            data_res = pd.DataFrame(results[var])
            data_res = data_res.drop(['Pol'], axis=1)
            data_res = data_res.drop(1)
            data_res = data_res.set_index('BC_Name')
            append_df_to_excel(resfile, data_res, core, startrow=2,
                                header=False, sheet_name=var[:3])
    else:
        for var in results:
            data_res = pd.DataFrame(results[var])
            data_res = data_res.drop(['Pol'], axis=1)
            data_res = data_res.drop(1)
            data_res = data_res.set_index('BC_Name')
            append_df_to_excel(resfile, data_res, core, header=False,
                               sheet_name=var[:3])
